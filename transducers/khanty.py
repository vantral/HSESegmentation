# -*- coding: utf-8 -*-
"""Khanty example.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/14PpMXNaBx-nVSD8PVwjTF6UI5ikMGjAc

Порядок слов в хантыйских существительных: 

**NounStem (NounMorph?) (NounNum?) (NounPoss?) (NounCase?)**, причём только в таком порядке

Предполагаемый алгоритм действий:

1) Спросить про обязательность категорий

2) Сделать PATTERNS из считанных данных (как перевести порядок морфем в PATTERNS? )

3) Сделать лексикон

4) Подтянуть словарь
"""

import re
from openpyxl import load_workbook
book = load_workbook(filename = "Морфология.xlsx")
sheet = book['Noun']
size = sheet.max_row
n = 'None'
#print(size)

#откроем файл в режиме чтения и записи (запись в конец файла)
file = open('khantynoun.lexd', 'a+', encoding = 'utf-8')
#вручную записываем блок существующих морфологических паттернов:
file.write("PATTERNS" + '\n')
file.write("NounStem (NounMorph?) (NounNum?) (NounPoss?) (NounCase?)" + '\n')
#идём по строкам Excel-файла, 
for i in range (2, size):
  if str(sheet['A' + str(i)].value) != 'None':
    if str(sheet['B' + str(i)].value) != 'None': #если в обеих колонках ячейки непустые
       if str(sheet['A' + str(i)].value) == 'LEXICON':   #и если в первой колонке стоит слово LEXICON
         file.write(str(sheet['A' + str(i)].value) +  ' ')  
         file.write(str(sheet['B' + str(i)].value) + '\n')   #то записываем в файл название лексикона
       else:
          file.write('<' + str(sheet['A' + str(i)].value) + '>' + ':')   
          file.write(str(sheet['B' + str(i)].value) + '\n')     #иначе пишем название категории в <>, : и саму морфему
file.close()

#записываем словарь:
with open("Словарь сущ без композитов.txt","r") as input:
    with open('khantynoun.lexd' ,"a+") as output:  #открываем файл, в котором уже есть морфемы и PATTERNS
       output.write("LEXICON NounStem\n")   #вручную записываем, что сейчас вводится лексикон с основами сущ
       output.write('/([a-zχəšśλŋńäɵă])+/'+ '\n')   #регулярка, которая помогает ловить несловарные корни
       for line in input:
              line = line.replace(" 			 ","#")   #т.к. словарь кривой, надо убрать некоторое кол-во пробелов из файла
              line = line.replace(" 			", "#")
              line = line.replace(" 	 ", "#")
#              if "#" not in line:
#                 line = line.replace(" ", "#", 1)
              output.write(line)
       output.write(' ')

  #print(sheet['A' + str(i)].value,sheet['B' + str(i)].value )

!curl -sS https://apertium.projectjj.com/apt/install-nightly.sh | sudo bash
!apt install apertium-all-dev lexd

!lexd khantynoun.lexd > noun-generator.att
!lt-comp rl noun-generator.att noun-analyser.bin

"""Мини-тест:"""

# Commented out IPython magic to ensure Python compatibility.
# %%writefile words.txt
# wojnaja 
# sowχosən 
# sowχosa 
# pensijaja 
# wošən 
# śaśiλam 
# aŋkśaśiλam 
# taśəλ 
# gotən 
# wʉλiλaλ 
# akɛm 
# taśa 
# utλaλ 
# aśɛm 
# ripaka
# akew 
# mɵša 
# χăntet 
# jasŋewən
# χotenən
# imeŋən

!cat words.txt | lt-proc noun-analyser.bin